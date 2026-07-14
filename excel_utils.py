from openpyxl import load_workbook
class ExcelUtils:

    @staticmethod
    def get_row_count(path, sheet_name):
        workbook = load_workbook(path)
        sheet = workbook[sheet_name]
        return sheet.max_row

    @staticmethod
    def read_data(path, sheet_name, row, column):
        workbook = load_workbook(path)
        sheet = workbook[sheet_name]
        return sheet.cell(row=row, column=column).value

    @staticmethod
    def write_data(path, sheet_name, row, column, data):
        workbook = load_workbook(path)
        sheet = workbook[sheet_name]
        sheet.cell(row=row, column=column).value = data
        workbook.save(path)